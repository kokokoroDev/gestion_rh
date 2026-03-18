from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

wb = Workbook()
ws = wb.active
ws.title = "Teletravail"

headers = ["Date", "Employé", "Heures", "Projet", "Tâche", "Statut", "Lieu", "Notes"]
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for col, header in enumerate(headers, 1):
    cell = ws.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')
    cell.border = thin_border

sample_data = [
    ["2026-03-16", "Dupont Jean", 8, "Projet A", "Développement", "Terminé", "Domicile", "OK"],
    ["2026-03-16", "Martin Sophie", 6, "Projet B", "Réunion", "En cours", "Domicile", ""],
    ["2026-03-17", "Dupont Jean", 8, "Projet A", "Tests", "Terminé", "Domicile", "OK"],
    ["2026-03-17", "Bernard Lucas", 7, "Projet C", "Analyse", "En cours", "Bureau", ""],
    ["2026-03-18", "Martin Sophie", 8, "Projet B", "Développement", "Terminé", "Domicile", "OK"],
    ["2026-03-18", "Dupont Jean", 4, "Projet A", "Réunion", "En cours", "Domicile", "Réunion d'équipe"],
    ["2026-03-19", "Bernard Lucas", 8, "Projet C", "Développement", "Terminé", "Bureau", "OK"],
    ["2026-03-19", "Martin Sophie", 8, "Projet B", "Documentation", "Terminé", "Domicile", "OK"],
]

for row_idx, row_data in enumerate(sample_data, 2):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

ws.column_dimensions['A'].width = 12
ws.column_dimensions['B'].width = 15
ws.column_dimensions['C'].width = 8
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 10
ws.column_dimensions['H'].width = 20

wb.save("teletravail.xlsx")
print("Fichier Excel créé: teletravail.xlsx")
